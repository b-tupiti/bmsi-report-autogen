import datetime
from openpyxl.styles import colors, fills, PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
from openpyxl.utils import get_column_letter
import numpy as np
def customize_sales(ws):

    ws.delete_cols(1)
    ws.delete_rows(2)
    
    _len = len(ws['A'])
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='e38181'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    ws['B1'].fill = fill
    ws['B1'].font = ft
    
    #Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'AB'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    return ws

def customize_fca(ws,_len):
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)

    #get month & Year
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year

    mon_year = str(month) + ' ' + str(year)


    ws.merge_cells('A1:B1')
    top_left_cell = ws['A1']
    top_left_cell.value = 'First Call Activation (FCA) - ' + mon_year;
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    
    last_cell = ws['B'+str(_len)]
    total_cell = ws['B'+str(_len+1)]
    tot_title_cell = ws['A'+str(_len+1)]

    last_cell_str = str(last_cell)
    last_cell_str = last_cell_str.split('.')[1][:-1]


    formula = "=SUM(B3:" + last_cell_str + ")"
    total_cell.value = formula
    tot_title_cell.value = "Total"

    for cell in ws['A']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws['B']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Colors 

    # head
    headfill_col = colors.Color(rgb='31859C')
    head_fill = fills.PatternFill(patternType='solid', fgColor=headfill_col)
    ws['A2'].fill = head_fill
    ws['B2'].fill = head_fill
    top_left_cell.fill = head_fill

    # total
    totfill_col = colors.Color(rgb='DDEBF7')
    tot_fill = fills.PatternFill(patternType='solid', fgColor=totfill_col)
    tot_title_cell.fill = tot_fill
    total_cell.fill = tot_fill

    #Fonts
    head_ft = Font(color=colors.WHITE,size=10,bold=True)
    ws['A2'].font = head_ft
    ws['B2'].font = head_ft
    top_left_cell.font = head_ft
    bottom_ft = Font(bold=True)
    tot_title_cell.font = bottom_ft
    total_cell.font = bottom_ft

    #Borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for cell in ws['A']:
        cell.border = thin_border
    for cell in ws['B']:
        cell.border = thin_border
        
    return ws

def customize_ass(ws,_len):
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)

    #get month & Year
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year

    mon_year = str(month) + ' ' + str(year)

    ws.merge_cells('A1:B1')
    top_left_cell = ws['A1']
    top_left_cell.value = 'Active Subscribers Status - ' + mon_year;
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20


    for cell in ws['A']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws['B']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Colors 

    # head
    headfill_col = colors.Color(rgb='31859C')
    head_fill = fills.PatternFill(patternType='solid', fgColor=headfill_col)
    ws['A2'].fill = head_fill
    ws['B2'].fill = head_fill
    top_left_cell.fill = head_fill

    #Fonts
    head_ft = Font(color=colors.WHITE,size=10,bold=True)
    ws['A2'].font = head_ft
    ws['B2'].font = head_ft
    top_left_cell.font = head_ft

    #Borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for cell in ws['A']:
        cell.border = thin_border
    for cell in ws['B']:
        cell.border = thin_border
    
    return ws

def customize_cdb(ws):
    
    REV_SUMMARY_COL = []
    
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    _len = len(ws['B'])
    
    #get month & Year
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year

    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:S1')
    title_tl_cell = ws['A1']
    title_tl_cell.value = 'Daily Call Detail Summary(Billable Calls) - ' + mon_year;
    title_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('B2:D2')
    on_tl_cell = ws['B2']
    on_tl_cell.value = 'ON NET'
    on_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('E2:G2')
    off_tl_cell = ws['E2']
    off_tl_cell.value = 'OFF NET'
    off_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('H2:J2')
    tel_tl_cell = ws['H2']
    tel_tl_cell.value = 'TELIKOM(PSTN)'
    tel_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('K2:M2')
    int_tl_cell = ws['K2']
    int_tl_cell.value = 'INTERNATIONAL'
    int_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('N2:P2')
    oth_tl_cell = ws['N2']
    oth_tl_cell.value = 'OTHERS'
    oth_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('Q2:S2')
    tot_tl_cell = ws['Q2']
    tot_tl_cell.value = 'TOTAL'
    tot_tl_cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws['Q3'].value = 'Nbr of Calls'
    ws['Q3'].alignment = Alignment(horizontal="center",vertical="center")
    
    ws['R3'].value = 'Minutes'
    ws['R3'].alignment = Alignment(horizontal="center",vertical="center")
    
    ws['S3'].value = 'Face Value($)'
    ws['S3'].alignment = Alignment(horizontal="center",vertical="center")
    
    
    #--------------------
    
    tot_title_cell = ws['A'+str(_len+1)]
    tot_title_cell.value = "Total"
    
    x = 'BCDEFGHIJKLMNOPQRS'
    
    for i in x:
        last_cell = ws[i+str(_len)]
        total_cell = ws[i+str(_len+1)]
        last_cell_str = str(last_cell)
        last_cell_str = last_cell_str.split('.')[1][:-1]
        formula = "=SUM(" + i + "4:" + last_cell_str + ")"
        total_cell.value = formula
        
   
     #---------------------
    
    x = 'ABCDEFGHIJKLMNOPQRS'
    
    for i in x:
        ws.column_dimensions[i].width = 16
        for cell in ws[i]:
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    _len = len(ws['B'])
    for i in range(4,_len):
        # Nbr of Calls
        formula = '=SUM(B' + str(i) + ',E' + str(i) + ',H' + str(i)+ ',K' + str(i) + ',N' + str(i) + ')'
        ws['Q'+str(i)].value = formula
        ws['Q'+str(i)].font = Font(color=colors.BLACK,size=10,bold=True)
        # Minutes
        formula = '=SUM(C' + str(i) + ',F' + str(i) + ',I' + str(i)+ ',L' + str(i) + ',O' + str(i) + ')'
        ws['R'+str(i)].value = formula
        ws['R'+str(i)].font = Font(color=colors.BLACK,size=10,bold=True)
        # Face Value($)
        formula = '=SUM(D' + str(i) + ',G' + str(i) + ',J' + str(i)+ ',M' + str(i) + ',P' + str(i) + ')'
        ws['S'+str(i)].value = formula
        ws['S'+str(i)].font = Font(color=colors.BLACK,size=10,bold=True)
        
        val = ws['D'+str(i)].value + ws['G'+str(i)].value + ws['J'+str(i)].value + ws['M'+str(i)].value + ws['P'+str(i)].value

        REV_SUMMARY_COL.append(val)
        
    #---------------------
    # Colors & Fonts
    
    head_ft = Font(color=colors.WHITE,size=10,bold=True)

    # head
    headfill_col = colors.Color(rgb='31859C')
    head_fill = fills.PatternFill(patternType='solid', fgColor=headfill_col)
    # total
    totl_col = colors.Color(rgb='DDEBF7')
    totl_fill = fills.PatternFill(patternType='solid', fgColor=totl_col)
    
    title_tl_cell.fill = head_fill
    title_tl_cell.font = head_ft
    
    on_tl_cell.fill = head_fill 
    on_tl_cell.font = head_ft
    
    off_tl_cell.fill = head_fill
    off_tl_cell.font = head_ft
    
    tel_tl_cell.fill = head_fill
    tel_tl_cell.font = head_ft
    
    int_tl_cell.fill = head_fill
    int_tl_cell.font = head_ft
    
    oth_tl_cell.fill = head_fill
    oth_tl_cell.font = head_ft
    
    tot_tl_cell.fill = head_fill
    tot_tl_cell.font = head_ft

    ws['A2'].fill = head_fill
    ws['A2'].font = head_ft
    
    ##---------
    
    for i in x:
        ws[i+'3'].fill = head_fill
        ws[i+'3'].font = head_ft
        ws[i+str(_len)].fill = totl_fill
    
    #---------------------
    
    for i in range(4,_len):
        ws['Q'+str(i)].fill = totl_fill
        ws['R'+str(i)].fill = totl_fill
        ws['S'+str(i)].fill = totl_fill
        
    #Fonts
    btm_ft = Font(color=colors.BLACK,size=10,bold=True)
    
    #Borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for i in x:
        ws[i+str(_len)].font = btm_ft
        for cell in ws[i]:
            cell.border = thin_border
    
    return REV_SUMMARY_COL

def customize_cdnb(ws):
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    # 2. Insert totals at bottom, and totals on the right side
    
    # bottom
    _len = len(ws['B'])
    ws['A' + str(_len + 1)].value = 'Total'
    
    x = 'BCDEFGHIJKLM'
    
    for i in x:
        ws[i+str(_len + 1)].value = '=SUM(' + i + '4:' + i + str(_len) + ')'
    
    # right-side
    ws['L3'].value = 'Nbr of Calls'
    ws['M3'].value = 'Minutes'
    
    for i in range(4,_len + 1):
        formula = '=SUM(B' + str(i) + ',D' + str(i) + ',F' + str(i)+ ',H' + str(i) + ',J' + str(i) + ')'
        ws['L'+str(i)].value = formula
        formula = '=SUM(C' + str(i) + ',E' + str(i) + ',G' + str(i)+ ',I' + str(i) + ',K' + str(i) + ')'
        ws['M'+str(i)].value = formula
        
    # 3. Insert and align merged Header Titles 
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:M1')
    ws.merge_cells('B2:C2')
    ws.merge_cells('D2:E2')
    ws.merge_cells('F2:G2')
    ws.merge_cells('H2:I2')
    ws.merge_cells('J2:K2')
    ws.merge_cells('L2:M2')
    
    ws['A1'] = 'Daily Call Detail Summary(Non Billable Calls) - ' + mon_year;
    ws['B2'].value = 'ON NET'
    ws['D2'].value = 'OFF NET'
    ws['F2'].value = 'TELIKOM (PSTN)'
    ws['H2'].value = 'INTERNATIONAL'
    ws['J2'].value = 'OTHERS'
    ws['L2'].value = 'TOTAL'
    
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    ws['A2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['B2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['D2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['F2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['H2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['J2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['L2'].alignment = Alignment(horizontal="center",vertical="center")
    
    # 4. Align all cells
    x = 'ABCDEFGHIJKLM'
    for i in range(3,_len+2):
        for _x in x:
            ws[_x+str(i)].alignment = Alignment(horizontal="center",vertical="center")

    # 5. Color and fonts for headers and totals
    
    # titles
    head_fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='31859C'))
    
    ws['A1'].fill = head_fill
    ws['A2'].fill = head_fill
    ws['B2'].fill = head_fill
    ws['D2'].fill = head_fill
    ws['F2'].fill = head_fill
    ws['H2'].fill = head_fill
    ws['J2'].fill = head_fill
    ws['L2'].fill = head_fill
    
    for i in x:
        ws[i+'3'].fill = head_fill
        
    # totals - RIGHT SIDE
    totl_fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='DDEBF7'))
    
    tot_ft = Font(color=colors.BLACK,size=10,bold=True)
    
    x = 'LM'
    for i in range(4,_len+1):
        for _x in x:
            ws[_x+str(i)].fill = totl_fill
            ws[_x+str(i)].font = tot_ft
        
    # totals - BOTTOM
    x = 'ABCDEFGHIJKLM'
    for i in x:
        ws[i+str(_len+1)].fill = totl_fill
        ws[i+str(_len+1)].font = tot_ft
    
    # 6. column dimensions and borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for i in x:
        ws.column_dimensions[i].width = 16
        for cell in ws[i]:
            cell.border = thin_border
    
    
    return ws

def customize_smsdb(ws):
    
    REV_COL = []
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    # 2. Insert totals at bottom, and totals on the right side
    
    # bottom
    _len = len(ws['B'])
    ws['A' + str(_len + 1)].value = 'Total'
    
    x = 'BCDEFGHIJK'
    
    for i in x:
        ws[i+str(_len + 1)].value = '=SUM(' + i + '4:' + i + str(_len) + ')'
    
    # right-side
    ws['J3'].value = 'Nbr of SMS'
    ws['K3'].value = 'Face Value($)'
    
    for i in range(4,_len + 1):
        # Nbr of SMS
        formula = '=SUM(B' + str(i) + ',D' + str(i) + ',F' + str(i)+ ',H' + str(i) + ')'
        ws['J'+str(i)].value = formula
        # Face Value
        formula = '=SUM(C' + str(i) + ',E' + str(i) + ',G' + str(i)+ ',I' + str(i) + ')'
        ws['K'+str(i)].value = formula
        val = ws['C'+str(i)].value + ws['E'+str(i)].value + ws['G'+str(i)].value + ws['I'+str(i)].value
        REV_COL.append(val)
    
    # 3. Insert and align merged Header Titles 
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:K1')
    ws.merge_cells('B2:C2')
    ws.merge_cells('D2:E2')
    ws.merge_cells('F2:G2')
    ws.merge_cells('H2:I2')
    ws.merge_cells('J2:K2')
    
    ws['A1'] = 'Summary (Billable) SMS - ' + mon_year;
    ws['B2'].value = 'ON NET'
    ws['D2'].value = 'OFF NET'
    ws['F2'].value = 'INTERNATIONAL'
    ws['H2'].value = 'OTHERS'
    ws['J2'].value = 'TOTAL'
    
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    ws['A2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['B2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['D2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['F2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['H2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['J2'].alignment = Alignment(horizontal="center",vertical="center")
    
    
    # 4. Align all cells
    x = 'ABCDEFGHIJK'
    for i in range(3,_len+2):
        for _x in x:
            ws[_x+str(i)].alignment = Alignment(horizontal="center",vertical="center")
    
    
    
    # 5. column dimensions and borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for i in x:
        ws.column_dimensions[i].width = 16
        for cell in ws[i]:
            cell.border = thin_border
    
    # 6. Color and fonts for headers and totals
    
    # titles
    head_fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='31859C'))
    
    ws['A1'].fill = head_fill
    ws['A2'].fill = head_fill
    ws['B2'].fill = head_fill
    ws['D2'].fill = head_fill
    ws['F2'].fill = head_fill
    ws['H2'].fill = head_fill
    ws['J2'].fill = head_fill
    
    for i in x:
        ws[i+'3'].fill = head_fill
        
    # totals - RIGHT SIDE
    totl_fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='DDEBF7'))
    
    tot_ft = Font(color=colors.BLACK,size=10,bold=True)
    
    x = 'JK'
    for i in range(4,_len+1):
        for _x in x:
            ws[_x+str(i)].fill = totl_fill
            ws[_x+str(i)].font = tot_ft
        
    # totals - BOTTOM
    x = 'ABCDEFGHIJK'
    for i in x:
        ws[i+str(_len+1)].fill = totl_fill
        ws[i+str(_len+1)].font = tot_ft
    
    #--------------------------------------------
    
    return REV_COL

def customize_smsdnb(ws):
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    
    # 2. Insert totals at bottom, and totals on the right side
    
    # bottom
    _len = len(ws['B'])
    ws['A' + str(_len + 1)].value = 'Total'
    
    x = 'BCDEF'
    
    for i in x:
        ws[i+str(_len + 1)].value = '=SUM(' + i + '3:' + i + str(_len) + ')'
    
    # right-side
    ws['F2'].value = 'Total SMS'
    
    for i in range(3,_len + 1):
        formula = '=SUM(B' + str(i) + ',C' + str(i) + ',D' + str(i)+ ',E' + str(i) + ')'
        ws['F'+str(i)].value = formula
        
    
    # 3. Insert and align merged Header Titles 
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Non Billable SMS Detail Summary - ' + mon_year;
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    
    # 4. Align all cells
    x = 'ABCDEF'
    for i in range(2,_len+2):
        for _x in x:
            ws[_x+str(i)].alignment = Alignment(horizontal="center",vertical="center")
    
    # 5. column dimensions and borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for i in x:
        ws.column_dimensions[i].width = 16
        for cell in ws[i]:
            cell.border = thin_border
    
    # 6. Color and fonts for headers and totals
    
    # titles
    head_fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='31859C'))
    
    ws['A1'].fill = head_fill
    for i in x:
        ws[i+'2'].fill = head_fill
        
    # totals - RIGHT SIDE
    totl_fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='DDEBF7'))
    
    tot_ft = Font(color=colors.BLACK,size=10,bold=True)
    
    x = 'F'
    for i in range(3,_len+1):
        for _x in x:
            ws[_x+str(i)].fill = totl_fill
            ws[_x+str(i)].font = tot_ft
        
    # totals - BOTTOM
    x = 'ABCDEF'
    for i in x:
        ws[i+str(_len+1)].fill = totl_fill
        ws[i+str(_len+1)].font = tot_ft
    
    #------------------
    return ws

def customize_gprsbnb(ws):
    
    rev_col = []
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    # 2. Insert totals at bottom
    
    
    _len = len(ws['B']) # length of each column
    ws['A' + str(_len + 1)].value = 'Total' # writing 'Total' to first cell of totals 
    
    # Columns to put totals
    x = 'BCDEFGHI'
    
    # putting formulas in the totals row cells
    for i in x:
        ws[i+str(_len + 1)].value = '=SUM(' + i + '4:' + i + str(_len) + ')'
    
    # 3. Insert and align merged Header Titles 
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Data Traffic - ' + mon_year
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('B2:E2')
    ws['B2'] = 'BILLABLE'
    ws['B2'].alignment = Alignment(horizontal="center",vertical="center")
    
    ws.merge_cells('F2:I2')
    ws['F2'] = 'NON BILLABLE'
    ws['F2'].alignment = Alignment(horizontal="center",vertical="center")
    
    # 4. Align all cells
    x = 'ABCDEFGHI'
    for i in range(3,_len+2):
        for _x in x:
            ws[_x+str(i)].alignment = Alignment(horizontal="center",vertical="center")
    
    # 5. column dimensions and borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for i in x:
        ws.column_dimensions[i].width = 16
        for cell in ws[i]:
            cell.border = thin_border
            
    # 6. Color and fonts for headers and totals
    
    # titles
    head_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='31859C'))
    
    ws['A1'].fill = head_fill
    ws['A2'].fill = head_fill
    ws['B2'].fill = head_fill
    ws['F2'].fill = head_fill
    for i in x:
        ws[i+'3'].fill = head_fill
        

    
        
    # totals - BOTTOM
    totl_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='DDEBF7'))
    
    tot_ft = Font(color=colors.BLACK,size=10,bold=True)
    x = 'ABCDEFGHI'
    for i in x:
        ws[i+str(_len+1)].fill = totl_fill
        ws[i+str(_len+1)].font = tot_ft
            
    return ws

def customize_sd(ws):
    
    rev_col = []
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    
    # 2. Insert and align merged Header Titles 
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:BA1')
    ws.merge_cells('B2:D2')
    ws.merge_cells('E2:G2')
    ws.merge_cells('H2:J2')
    ws.merge_cells('K2:M2')
    ws.merge_cells('N2:P2')
    ws.merge_cells('Q2:S2')
    ws.merge_cells('T2:V2')
    ws.merge_cells('W2:Y2')
    ws.merge_cells('Z2:AB2')
    ws.merge_cells('AC2:AE2')
    ws.merge_cells('AF2:AH2')
    ws.merge_cells('AI2:AK2')
    ws.merge_cells('AL2:AN2')
    ws.merge_cells('AO2:AQ2')
    ws.merge_cells('AR2:AT2')
    ws.merge_cells('AU2:AW2')
    ws.merge_cells('AX2:AZ2')
    ws.merge_cells('A2:A3')
    ws.merge_cells('BA2:BA3')
    
    ws['A1'] = 'SUBSCRIPTION BUNDLES - ' + mon_year
    ws['A2'] = 'DATE'
    ws['BA2'] = 'Total Amount'
    ws['B2'] = 'Global'
    ws['E2'] = 'Moa Day ($7)'
    ws['H2'] = 'Moa 2 Days ($14)'
    ws['K2'] = 'Moa 3 Days ($21)'
    ws['N2'] = 'Moa Week ($42)'
    ws['Q2'] = 'Moa Month ($500)'
    ws['T2'] = 'D6 ($6)'
    ws['W2'] = 'Hour Data ($10)'
    ws['Z2'] = 'D15 ($15)'
    ws['AC2'] = 'D2GB ($50)'
    ws['AF2'] = 'D20 ($20)'
    ws['AI2'] = 'Movie Night ($35)'
    ws['AL2'] = 'Week Data ($50)'
    ws['AO2'] = 'D90 ($90)'
    ws['AR2'] = 'D220 ($220)'
    ws['AU2'] = 'D500 ($500)'
    ws['AX2'] = 'Roaming Bundle'

    
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    ws['A2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['B2'].alignment = Alignment(horizontal="center",vertical="center") 
    ws['E2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['H2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['K2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['N2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['Q2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['T2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['W2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['Z2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AC2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AF2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AI2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AL2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AO2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AR2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AU2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['AX2'].alignment = Alignment(horizontal="center",vertical="center")
    ws['BA2'].alignment = Alignment(horizontal="center",vertical="center")
    
    # 3. Insert totals
    
    
    _len = len(ws['B']) # length of each column
    ws['A' + str(_len + 1)].value = 'Total' # writing 'Total' to first cell of totals 
    
    # Columns to put totals
    # x = 'BCDEFGHI'
    x = ['B','C','D','E','F','G','H','I','J','K',
         'L','M','N','O','P','Q','R','S','T','U',
         'V','W','X','Y','Z','AA','AB','AC','AD',
         'AE','AF','AG','AH','AI','AJ','AK','AL',
         'AM','AN','AO','AP','AQ','AR','AS','AT',
         'AU','AV','AW','AX','AY','AZ','BA']
    
    # putting formulas in the totals row cells
    for i in x:
        ws[i+str(_len + 1)].value = '=SUM(' + i + '4:' + i + str(_len) + ')'
    
    # RIGHT side
    for i in range(4,_len + 1):
        formula = '=SUM(C' + str(i) + ',F' + str(i) + ',I' + str(i)+ ',L' + str(i) + ',O'+ str(i) + ',R'+ str(i) + ',U'+ str(i) + ',X' + str(i) +',AA'+ str(i) + ',AD'+ str(i) + ',AG'+ str(i) + ',AJ' + str(i) +',AM'+ str(i) + ',AP' + str(i) +',AS' + str(i) +',AV' + str(i) + ',AY' +str(i)+ ')' 
        ws['BA'+str(i)].value = formula
        val = ws['C'+str(i)].value + ws['F'+str(i)].value + ws['I'+str(i)].value + ws['L'+str(i)].value + ws['O'+str(i)].value + ws['R'+str(i)].value + ws['U'+str(i)].value + ws['X'+str(i)].value + ws['AA'+str(i)].value + ws['AD'+str(i)].value + ws['AG'+str(i)].value + ws['AJ'+str(i)].value + ws['AM'+str(i)].value + ws['AP'+str(i)].value + ws['AS'+str(i)].value + ws['AV'+str(i)].value + ws['AY'+str(i)].value
        rev_col.append(val)
        
    
    
    # 4. Align all cells
    x = ['A','B','C','D','E','F','G','H','I','J','K',
         'L','M','N','O','P','Q','R','S','T','U',
         'V','W','X','Y','Z','AA','AB','AC','AD',
         'AE','AF','AG','AH','AI','AJ','AK','AL',
         'AM','AN','AO','AP','AQ','AR','AS','AT',
         'AU','AV','AW','AX','AY','AZ','BA']
    for i in range(4,_len+2):
        for _x in x:
            ws[_x+str(i)].alignment = Alignment(horizontal="center",vertical="center")
    
    # 5. column dimensions and borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 14
        for cell in ws[i]:
            cell.border = thin_border
    
    
        
    # 6. Color and fonts for headers and totals
    
    # Header
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='4a4d4f'))
    ft = Font(color=colors.WHITE,size=10,bold=True)
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    # Totals 
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='9c9797'))
    ft = Font(color=colors.WHITE,size=14,bold=True)
    for _x in x:
        ws[_x+str(_len+1)].fill = fill
        ws[_x+str(_len+1)].font = ft
    
    ws['BA2'].fill = fill
    ws['BA2'].font = ft
    for i in range(4,_len+1):
        ws['BA'+str(i)].fill = fill
        ws['BA'+str(i)].font = ft
    
    # Global
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='d1c252'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['B2'].fill = fill
    ws['B2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'BCD'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft

    # Moa Day ($7)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='7cd463'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['E2'].fill = fill
    ws['E2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'EFG'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # Moa 2 Days ($14)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='61b8ed'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['H2'].fill = fill
    ws['H2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'HIJ'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # Moa 3 Days ($21)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='cc99c7'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['K2'].fill = fill
    ws['K2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'KLM'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    # Moa Week ($42)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='ed9dbb'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['N2'].fill = fill
    ws['N2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'NOP'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft        
    
    # Moa Month ($500)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='dbedc7'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['Q2'].fill = fill
    ws['Q2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'QRS'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # D6 ($6)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='f5f19f'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['T2'].fill = fill
    ws['T2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'TUV'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    # Hour Data ($10)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='40870e'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['W2'].fill = fill
    ws['W2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = 'WXY'
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    # D15 ($15)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='e38181'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['Z2'].fill = fill
    ws['Z2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['Z','AA','AB']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    # D2GB ($50)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='6b7eed'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AC2'].fill = fill
    ws['AC2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AC','AD','AE']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    # D20 ($20)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='d16d6d'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AF2'].fill = fill
    ws['AF2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AF','AG','AH']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft        
    
    # Movie Night ($35)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='d3d4b6'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AI2'].fill = fill
    ws['AI2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AI','AJ','AK']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # Week Data ($50)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='ff9ead'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AL2'].fill = fill
    ws['AL2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AL','AM','AN']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # D90 ($90)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='939460'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AO2'].fill = fill
    ws['AO2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AO','AP','AQ']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # D220 ($220)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='8150eb'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AR2'].fill = fill
    ws['AR2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AR','AS','AT']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    # D500 ($500)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='9ff792'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AU2'].fill = fill
    ws['AU2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AU','AV','AW']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
    
    # Roaming Bundle
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='7de3d0'))
    ft = Font(color=colors.BLACK,size=12,bold=True)
    ws['AX2'].fill = fill
    ws['AX2'].font = ft    
    
    ft = Font(color=colors.BLACK,size=10,bold=False)
    x = ['AX','AY','AZ']
    for _x in x:
        for i in range(3,_len+1):
            ws[_x+str(i)].fill = fill
            ws[_x+str(i)].font = ft
            
    return (ws,rev_col)


def customize_rbs(ws):
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    
     # 2. Insert and align merged Header Titles 
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:D1')
    ws['A1'].value = 'Recharge by Subscribers - ' + mon_year
    
    # 3. Insert totals

    _len = len(ws['B']) # length of each column
    ws['A' + str(_len + 1)].value = 'Total' # writing 'Total' to first cell of totals 
    
    x = 'BCD'
    
    # putting formulas in the totals row cells
    for i in x:
        ws[i+str(_len + 1)].value = '=SUM(' + i + '3:' + i + str(_len) + ')'
    
    # 4. Borders and Alignment
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'ABCD'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
            
    # 5. Colors and fonts
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='8150eb'))
    ft = Font(color=colors.WHITE,size=11,bold=True)
    ws['A1'].fill = fill
    ws['A1'].font = ft   
    ft = Font(color=colors.WHITE,size=11,bold=False)
    for i in x:
        ws[i+'2'].fill = fill
        ws[i+'2'].font = ft
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='bfbaba'))
    ft = Font(color=colors.BLACK,size=11,bold=True)
    for i in x:
        ws[i+str(_len + 1)].fill = fill
        ws[i+str(_len + 1)].font = ft
        
    
    return ws

def customize_bt(ws):
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    
    # 2. Totals at the bottom
    _len = (len(ws['B']))
    x = 'BCDE'
    for i in x:
        ws[i+str(_len+1)].value = '=SUM(' + i + '3:' + i + str(_len) + ')'
    
    # 3. Merge and Texts
    ws.merge_cells('A1:E1')
    ws['A1'].value = 'Balance Transfer Activity -' + mon_year
    ws['A'+str(_len + 1)].value = 'Total'
    
    # 4. Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'ABCDE'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
            
    # 5. Styles
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='66a80f'))
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    ws['A1'].fill = fill
    ws['A1'].font = ft   
    for i in x:
        ws[i+'2'].fill = fill
        ws[i+'2'].font = ft
        
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='99e339'))
    ft = Font(color=colors.BLACK,size=11,bold=True)
    
    x = 'BCDE'
    for i in x:
        ws[i+str(_len + 1)].fill = fill
        ws[i+str(_len + 1)].font = ft
    
    
    #------------------------
    return ws

def customize_rm(ws):
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    # 2. Totals at the bottom
    _len = (len(ws['B']))
    x = 'BCDEFGHIJKLMNOPQRST'
    for i in x:
        ws[i+str(_len+1)].value = '=SUM(' + i + '4:' + i + str(_len) + ')'
    
    # 3. Totals on the right
    for i in range(4,_len+1):
        i_str = str(i)
        formula = '=SUM(C' + i_str + ',F' + i_str + ',I' + i_str + ',L' + i_str + ',O'+ i_str + ',R' + i_str + ')'
        ws['T' + i_str].value = formula
            
    
    # 4. Merge and Texts
    ws.merge_cells('A1:T1')
    ws['A1'].value = 'Balance Transfer Activity -' + mon_year
    ws['A'+str(_len + 1)].value = 'Total'
    
    ws.merge_cells('T2:T3')
    ws['T2'].value = 'Total'
    
    ws.merge_cells('A2:A3')
    ws['A2'].value = 'Date'
    
    ws.merge_cells('B2:D2')
    ws['B2'].value = 'EVD'
    
    ws.merge_cells('E2:G2')
    ws['E2'].value = 'cctopup'
    
    ws.merge_cells('H2:J2')
    ws['H2'].value = 'BSP'
    
    ws.merge_cells('K2:M2')
    ws['K2'].value = 'Staff Top Up'
    
    ws.merge_cells('N2:P2')
    ws['N2'].value = 'CUG/DUG'
    
    ws.merge_cells('Q2:S2')
    ws['Q2'].value = 'POB'
    
    
    # 4. Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'ABCDEFGHIJKLMNOPQRST'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    
    # 5. Styles
    
    # header
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='595e51'))
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    # EVD
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='c77367'))
    
    
    ws['B2'].fill = fill
    ws['B2'].font = ft 
    
    ws['B3'].fill = fill
    ws['B3'].font = ft
    
    ws['C3'].fill = fill
    ws['C3'].font = ft
    
    ws['D3'].fill = fill
    ws['D3'].font = ft
    
    # cctopup
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='00494d'))
    
    ws['E2'].fill = fill
    ws['E2'].font = ft 
    
    ws['E3'].fill = fill
    ws['E3'].font = ft
    
    ws['F3'].fill = fill
    ws['F3'].font = ft
    
    ws['G3'].fill = fill
    ws['G3'].font = ft
    
    # BSP
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='787005'))
    
    ws['H2'].fill = fill
    ws['H2'].font = ft 
    
    ws['H3'].fill = fill
    ws['H3'].font = ft
    
    ws['I3'].fill = fill
    ws['I3'].font = ft
    
    ws['J3'].fill = fill
    ws['J3'].font = ft
    
    # Staff Top Up
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='66a80f'))
    
    ws['K2'].fill = fill
    ws['K2'].font = ft 
    
    ws['K3'].fill = fill
    ws['K3'].font = ft
    
    ws['L3'].fill = fill
    ws['L3'].font = ft
    
    ws['M3'].fill = fill
    ws['M3'].font = ft
    
    # CUG/DUG 
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='beabf7'))
    
    ws['N2'].fill = fill
    ws['N2'].font = ft 
    
    ws['N3'].fill = fill
    ws['N3'].font = ft
    
    ws['O3'].fill = fill
    ws['O3'].font = ft
    
    ws['P3'].fill = fill
    ws['P3'].font = ft
    
    # POB
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='ff70a5'))
    
    ws['Q2'].fill = fill
    ws['Q2'].font = ft 
    
    ws['Q3'].fill = fill
    ws['Q3'].font = ft
    
    ws['R3'].fill = fill
    ws['R3'].font = ft
    
    ws['S3'].fill = fill
    ws['S3'].font = ft
    
    # Totals right side
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='edf1ff'))
    ft = Font(color=colors.BLACK,size=11,bold=True)
    
    for i in range(2,_len+1):
        ws['T'+str(i)].fill = fill
        ws['T'+str(i)].font = ft
    
    # Totals bottom
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='ffedf4'))
    ft = Font(color=colors.BLACK,size=11,bold=True)
    
    x = 'ABCDEFGHIJKLMNOPQRST'
    
    for i in x:
        ws[i+str(_len+1)].fill = fill
        ws[i+str(_len+1)].font = ft
        
    return ws

def customize_im(ws):
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    # 2. Totals at the bottom
    _len = (len(ws['B']))
    x = 'BCDEFG'
    for i in x:
        ws[i+str(_len+1)].value = '=SUM(' + i + '4:' + i + str(_len) + ')'
    
    # 3. Merges and Texts
    ws.merge_cells('A1:G1')
    ws['A1'].value = 'Incoming Call Minutes -' + mon_year
    ws['A'+str(_len + 1)].value = 'Total'
    
    ws.merge_cells('A2:A3')
    ws['A2'].value = 'DATE'
    
    ws.merge_cells('B2:C2')
    ws['B2'].value = 'Incomming Onnet'
    
    ws.merge_cells('D2:E2')
    ws['D2'].value = 'Incomming Offnet'
    
    ws.merge_cells('F2:G2')
    ws['F2'].value = 'Incomming PSTN'
    
    # 4. Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'ABCDEFG'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    # 5. Styles
    
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    # header
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='04596e'))
    
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    # Categories
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='065c22'))
    
    ws['B2'].fill = fill
    ws['B2'].font = ft
    ws['D2'].fill = fill
    ws['D2'].font = ft
    ws['F2'].fill = fill
    ws['F2'].font = ft
    
    # Col headers
    ft = Font(color=colors.WHITE,size=11,bold=False)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='69005e'))
    
    ws['A2'].fill = fill
    ws['A2'].font = ft
    
    x = 'BCDEFG'
    for i in x:
        ws[i+'3'].fill = fill
        ws[i+'3'].font = ft
    
    # Total bottom
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='eddaeb'))
    ft = Font(color=colors.BLACK,size=11,bold=True)
    
    x = 'ABCDEFG'
    
    for i in x:
        ws[i+str(_len+1)].fill = fill
        ws[i+str(_len+1)].font = ft
        
    return ws


def customize_err(ws):
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    
    # 2. Insert total at the bottom
    _len = len(ws['B'])
    ws['A'+str(_len+1)].value = 'TOTAL'
    ws['B'+str(_len+1)].value = '=SUM(B3:B' + str(_len) + ')'
    
    # 3. Merges and Texts
    ws.merge_cells('A1:B1')
    ws['A1'].value = 'Expired Resources Revenue -' + mon_year
    
    # 4. Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'AB'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 20
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    # Styles 
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='031536'))
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='083485'))
    ft = Font(color=colors.WHITE,size=12,bold=True)
    
    ws['A2'].fill = fill
    ws['A2'].font = ft
    ws['B2'].fill = fill
    ws['B2'].font = ft
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='5991f7'))
    ft = Font(color=colors.BLACK,size=11,bold=True)
    
    ws['A' + str(_len + 1)].fill = fill
    ws['A' + str(_len + 1)].font = ft
    ws['B' + str(_len + 1)].fill = fill
    ws['B' + str(_len + 1)].font = ft
    
    
    return ws


def customize_cp(ws):
    
    # Setup
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='5991f7'))
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(2)
    ws.delete_rows(2)
    ws['A1'].value = 'Entry Date'
    
    # 2. Count number of generated columns
    col_count = 0
    first_row = ws[1]
    for cell in first_row:
        if cell:
            cell.fill = fill
            cell.font = ft
            col_count += 1
     
    # 3. Count number of generated rows
    row_count = 0
    for cell in ws['A']:
        if cell.value is None:
            break
        row_count += 1
    
    # 4. Insert totals at the bottom
    for i in range(2,col_count + 2): # start at the second column and add an extra col
        col_letter = get_column_letter(i)
        ws[col_letter+str(row_count + 1)].fill = fill 
        ws[col_letter+str(row_count + 1)].font = ft
        ws[col_letter+str(row_count + 1)].value = '=SUM(' + col_letter + '2:' + col_letter + str(row_count) + ')'         
    
    # 5. Insert totals on the right side
    totals_column = get_column_letter(col_count+1)
    for i_row in range(2,row_count+1):
        
        # build formula for each total cell at the right
        formula = '=SUM('
        for i_col in range(2,col_count+1):
            cur_col = get_column_letter(i_col)
            formula += cur_col + str(i_row) + ','
        
        formula = formula[:-1]
        formula += ')'
        ws[totals_column+str(i_row)].value = formula
        ws[totals_column+str(i_row)].fill = fill
        ws[totals_column+str(i_row)].font = ft

    # 6. color total cells    
    ws['A'+str(row_count+1)].fill = fill
    ws['A'+str(row_count+1)].font = ft
    ws['A'+str(row_count+1)].value = 'Grand Total'  
    
    ws[get_column_letter(col_count+1)+'1'].fill = fill
    ws[get_column_letter(col_count+1)+'1'].font = ft
    ws[get_column_letter(col_count+1)+'1'].value = 'GrandTotal'
    
    # 7. Borders and Alignment
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for i in range(1,col_count+1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 17
        for cell in ws[col]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    return ws

def calc_gt_cp(ws):
    
    rev_col = []

    col_count = len(ws[1])
    row_count = len(ws['A'])

    for i_row in range(2,row_count):
        row_total = 0
        for i_col in range(2,col_count):
            if not np.isnan(ws[get_column_letter(i_col)+str(i_row)].value):
                row_total += ws[get_column_letter(i_col)+str(i_row)].value
        rev_col.append(row_total)

    return rev_col

def customize_rvbd(ws):
    
    # Setup
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='5991f7'))
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(2)
    ws.delete_rows(2)
    ws['A1'].value = 'Entry Date'
    
    # 2. Count number of generated columns
    col_count = 0
    first_row = ws[1]
    for cell in first_row:
        if cell:
            cell.fill = fill
            cell.font = ft
            col_count += 1
     
    # 3. Count number of generated rows
    row_count = 0
    for cell in ws['A']:
        if cell.value is None:
            break
        row_count += 1
    
    # 4. Insert totals at the bottom
    for i in range(2,col_count + 2): # start at the second column and add an extra col
        col_letter = get_column_letter(i)
        ws[col_letter+str(row_count + 1)].fill = fill 
        ws[col_letter+str(row_count + 1)].font = ft
        ws[col_letter+str(row_count + 1)].value = '=SUM(' + col_letter + '2:' + col_letter + str(row_count) + ')'         
    
    # 5. Insert totals on the right side
    totals_column = get_column_letter(col_count+1)
    for i_row in range(2,row_count+1):
        
        # build formula for each total cell at the right
        formula = '=SUM('
        for i_col in range(2,col_count+1):
            cur_col = get_column_letter(i_col)
            formula += cur_col + str(i_row) + ','
        formula = formula[:-1]
        formula += ')'
        ws[totals_column+str(i_row)].value = formula
        ws[totals_column+str(i_row)].fill = fill
        ws[totals_column+str(i_row)].font = ft

    # 6. color total cells    
    ws['A'+str(row_count+1)].fill = fill
    ws['A'+str(row_count+1)].font = ft
    ws['A'+str(row_count+1)].value = 'Grand Total'  
    
    ws[get_column_letter(col_count+1)+'1'].fill = fill
    ws[get_column_letter(col_count+1)+'1'].font = ft
    ws[get_column_letter(col_count+1)+'1'].value = 'GrandTotal'
    
    # 7. Borders and Alignment
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for i in range(1,col_count+1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 17
        for cell in ws[col]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    return ws

def customize_ocug_gsd(ws):
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    
    # 3. Merges and Texts
    ws.merge_cells('A1:E1')
    ws['A1'].value = 'Other CUG Gift Subscriptions - Deductions -' + mon_year
    
    
    # 4. Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'ABCDE'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    # 5. Styles
    
    ft = Font(color=colors.WHITE,size=11,bold=True)
    
    # header
    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='04596e'))
    
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    for i in x:
        ws[i+'2'].fill = fill
        ws[i+'2'].font = ft
    
    return ws

def customize_srev(ws):
    
    # 1. Delete and Insert rows and columns
    ws.delete_cols(1)
    ws.delete_rows(2)
    ws.insert_rows(1)
    
    # 2. Merge and texts
    
    from datetime import datetime
    month = datetime.now().strftime('%B')
    year = datetime.now().year
    mon_year = str(month) + ' ' + str(year)
    
    ws.merge_cells('A1:J1')
    ws['A1'] = 'SI Summary Revenue ' + mon_year
    
    
    
    # 2. Totals on the right
    _len = len(ws['A'])
    
    for i in range(3,_len+1):
        i_str = str(i)
        formula = '=SUM(B' + i_str + ',C' + i_str + ',D' + i_str + ',E' + i_str + ',F'+ i_str + ',G' + i_str + ',H' + i_str + ',I' + i_str + ')'
        ws['J' + i_str].value = formula
    ws['J2'].value = 'Total'
        
    # 3. Totals at the bottom
    _len = (len(ws['B']))
    x = 'BCDEFGHIJ'
    for i in x:
        ws[i+str(_len+1)].value = '=SUM(' + i + '3:' + i + str(_len) + ')'
    
    ws['A'+str(_len+1)].value = 'Total'
    
    # 4. Borders and Alignments
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    x = 'ABCDEFGHIJ'
    ws['A1'].border = thin_border
    for i in x:
        ws.column_dimensions[i].width = 17
        for cell in ws[i]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center",vertical="center")
    
    # Styles
    
    ft = Font(color=colors.BLACK,size=11,bold=True)    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='fffbb0'))
    for i in x:
        ws[i + str(_len+1)].fill = fill
        ws[i + str(_len+1)].font = ft
    
    for i in range(3,_len+2):
        ws['J' + str(i)].fill = fill
        ws['J' + str(i)].font = ft
    
    ws['J2'].fill = fill
    ws['J2'].font = ft
    
    #------
    ft = Font(color=colors.WHITE,size=11,bold=True)
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='221445'))
    
    ws['A1'].fill = fill
    ws['A1'].font = ft
    
    x = 'ABCDEFGHI'
    for i in x:
        ws[i+'2'].fill = fill
        ws[i+'2'].font = ft
        
    return ws