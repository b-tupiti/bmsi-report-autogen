from helper import listfilter, df_converter, customizer, grapher, generate_filename
import pandas as pd 
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def generate(sales_filename, report_dump_filename):
    
    # 1. get filenames
    dir_name = "input_files/"
    sales_file = dir_name + sales_filename
    report_dump_file = dir_name + report_dump_filename

    # 2. sales file
    sales = []
    with open(sales_file) as file:
        for line in file:
            sales.append(line)

    sales = sales[2:-1]
    sales_arr = []
    

    for i in range (len(sales)):
        sales[i] = sales[i].replace(" ","")
        sales[i] = sales[i].replace("\n", "")
        arr = sales[i].split(',')
        arr[0] = arr[0][:-8]
        arr[0] = arr[0].replace("/","-")
        row = arr[0] + '|' + arr[1]
        sales_arr.append(row)
    
    sales_arr = listfilter.filter_for_month(sales_arr)
    SALES_DF = df_converter.sales_to_df(sales_arr)
    
    # 3. report dump file

    with open(report_dump_file) as file:
        segment = []
        whole = []
        for line in file:
            # if line is not emp
            if len(line.strip()) == 0:
                whole.append(segment)
                segment = []
            else:
                segment.append(line) 

    for indx1 in range(len(whole)):
        whole[indx1] = whole[indx1][3:-1]
        for indx2 in range(len(whole[indx1])):
            whole[indx1][indx2] =  whole[indx1][indx2].replace("\n","")
            whole[indx1][indx2] =  whole[indx1][indx2].replace(" ","") 
    
    
    # Workbook 

    wb = Workbook()


    # Sales Sheet

    try:
        wb.create_sheet('Sales')
        ws = wb['Sales']
        
        rows = dataframe_to_rows(SALES_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws = customizer.customize_sales(ws)
    except:
        print('cannot generate Sales Sheet')
        pass

    # FCA Summary Sheet
    try:
        FCA_dl = listfilter.filter_for_month(whole[0]) 
        FCA_DF = df_converter.FCA_list_to_df(FCA_dl) 
        wb.create_sheet('FCA Summary')
        
        ws = wb['FCA Summary']

        rows = dataframe_to_rows(FCA_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        _len = len(ws['B'])
        ws = customizer.customize_fca(ws,_len)
        ws = grapher.graph_fca(ws,_len)
    except:
        print('cannot generate FCA Sheet')
        pass

    # Active Subs

    try:
        ASS_dl = listfilter.filter_for_month(whole[1]) 
        ASS_DF = df_converter.ASS_list_to_df(ASS_dl) 
        wb.create_sheet('Active Subs')
        
        ws = wb['Active Subs']

        rows = dataframe_to_rows(ASS_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        _len = len(ws['B'])
        ws = customizer.customize_ass(ws,_len)
        ws = grapher.graph_ass(ws,_len)
    except:
        print('cannot generate Active Subs Sheet')
        pass

    # Call Detail Billable
    try:
        CDB_dl = listfilter.filter_for_month(whole[2]) 
        CDB_DF = df_converter.CDB_list_to_df(CDB_dl) 
        wb.create_sheet('Call Detail Billable')
        
        ws = wb['Call Detail Billable']

        rows = dataframe_to_rows(CDB_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
 
        Voice_Billable = customizer.customize_cdb(ws)
    except:
        print('cannot generate Call Detail Billable Sheet')
        pass

    # Call Detail Non Billable

    try:
        CDNB_dl = listfilter.filter_for_month(whole[3]) 
        CDNB_DF = df_converter.CDNB_list_to_df(CDNB_dl) 
        wb.create_sheet('Call Detail Non Billable')
        
        ws = wb['Call Detail Non Billable']

        rows = dataframe_to_rows(CDNB_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws = customizer.customize_cdnb(ws)
    except:
        print('cannot generate Call Detail Non Billable Sheet')
        pass

    # SMS Billable

    try:
        SMSDB_dl = listfilter.filter_for_month(whole[4]) 
        SMSDB_DF = df_converter.SMSDB_list_to_df(SMSDB_dl) 
        wb.create_sheet('SMS Billable')
        
        ws = wb['SMS Billable']

        rows = dataframe_to_rows(SMSDB_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        SMS_Billable = customizer.customize_smsdb(ws)
    except:
        print('cannot generate SMS Billable Sheet')
        pass

    # SMS Non Billable

    try:
        SMSDNB_dl = listfilter.filter_for_month(whole[5]) 
        SMSDNB_DF = df_converter.SMSDNB_list_to_df(SMSDNB_dl) 
        wb.create_sheet('SMS Non Billable')
        
        ws = wb['SMS Non Billable']

        rows = dataframe_to_rows(SMSDNB_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws = customizer.customize_smsdnb(ws)
    except:
        print('cannot generate SMS Non Billable Sheet')
        pass

    # Data Billable Non Billable

    try:
        GPRSBNB_dl = listfilter.filter_for_month(whole[6]) 
        GPRSBNB_DF = df_converter.GPRSBNB_list_to_df(GPRSBNB_dl) 
        wb.create_sheet('Data Billable Non Billable')
        
        ws = wb['Data Billable Non Billable']

        rows = dataframe_to_rows(GPRSBNB_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws = customizer.customize_gprsbnb(ws)
        GPRS_Billable = []
        SR_Date = []
        for cell in ws['D'][3:-1]:
            GPRS_Billable.append(cell.value)
        for cell in ws['A'][3:-1]:
            SR_Date.append(cell.value)
    except:
        print('cannot generate Data Billable Non Billable Sheet')
        pass

    # Subscription Bundles

    try:
        SD_dl = listfilter.filter_for_month(whole[7])  
        SD_DF = df_converter.SD_list_to_df(SD_dl) 
        wb.create_sheet('Subscription Bundles')
        
        ws = wb['Subscription Bundles']

        rows = dataframe_to_rows(SD_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        _len = len(ws['B'])
        ws,Subscriptions = customizer.customize_sd(ws)
        ws = grapher.graph_sd(ws,_len)
    except:
        print('cannot generate Subscription Bundles Sheet')
        pass

    # Recharge by Subscribers

    try:
        RbS_dl = listfilter.filter_for_month(whole[8]) 
        RBS_DF = df_converter.RbS_list_to_df(RbS_dl) 
        wb.create_sheet('Recharge by Subscribers')
        
        ws = wb['Recharge by Subscribers']

        rows = dataframe_to_rows(RBS_DF)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        _len= len(ws['B'])
        ws = customizer.customize_rbs(ws)
        ws = grapher.graph_rbs(ws,_len)
    except:
        print('cannot generate Recharge by Subscribers Sheet')
        pass

    # Balance Transfer

    try:
        BT_dl = listfilter.filter_for_month(whole[9]) 
        BT_DF = df_converter.BT_list_to_df(BT_dl) 
        wb.create_sheet('Balance Transfer')
        
        ws = wb['Balance Transfer']

        rows = dataframe_to_rows(BT_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        ws = customizer.customize_bt(ws)
        Balance_Transfer = []
        for cell in ws['E'][2:-1]:
            Balance_Transfer.append(cell.value)
    except:
        print('cannot generate Balance Transfer Sheet')
        pass

    # Recharge Method

    try:
        RM_dl = listfilter.filter_for_month(whole[10]) 
        RM_DF = df_converter.RM_list_to_df(RM_dl) 
        wb.create_sheet('Recharge Method')
        
        ws = wb['Recharge Method']

        rows = dataframe_to_rows(RM_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        ws = customizer.customize_rm(ws)
    except:
        print('cannot generate Recharge Method Sheet')
        pass

    # Incoming Calls

    try:
        IM_dl = listfilter.filter_for_month(whole[11]) 
        IM_DF = df_converter.IM_list_to_df(IM_dl) 
        wb.create_sheet('Incoming Calls')
        
        ws = wb['Incoming Calls']

        rows = dataframe_to_rows(IM_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        _len= len(ws['B'])
        ws = customizer.customize_im(ws)
        ws = grapher.graph_im(ws,_len)
    except:
        print('cannot generate Incoming Calls Sheet')
        pass

    # Expired Resources Revenue

    try:
        ERR_dl = listfilter.filter_for_month(whole[12]) 
        ERR_DF = df_converter.ERR_list_to_df(ERR_dl) 
        wb.create_sheet('Expired Resources Revenue')
        
        ws = wb['Expired Resources Revenue']

        rows = dataframe_to_rows(ERR_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        ws = customizer.customize_err(ws)
        Expired_Revenue = []
        for cell in ws['B'][2:-1]:
            Expired_Revenue.append(cell.value)
    except:
        print('cannot generate Expired Resources Revenue')
        pass

    # Corporate Plan Revenue

    try:
        CP_dl = listfilter.filter_for_month(whole[14]) 
        CP_DF = df_converter.CP_list_to_df(CP_dl)    
        wb.create_sheet('Corporate Plan Revenue')
        
        ws = wb['Corporate Plan Revenue']

        CP_DF = CP_DF.pivot(index='Date',
                            columns='Product_Desc',
                            values='Revenue')

        rows = dataframe_to_rows(CP_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        ws = customizer.customize_cp(ws)
        Corporate_Plans = customizer.calc_gt_cp(ws)
    except:
        print('cannot generate Corporate Plan Revenue Sheet')
        pass

    # Other CUG Gift Subs Deductions

    try:
        oCUG_GSD_dl = listfilter.filter_for_month(whole[15]) 
        OCUG_GSD_DF = df_converter.oCUG_GSD_list_to_df(oCUG_GSD_dl) 
        wb.create_sheet('Other CUG Gift Subs Deductions')
        
        ws = wb['Other CUG Gift Subs Deductions']

        rows = dataframe_to_rows(OCUG_GSD_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        ws = customizer.customize_ocug_gsd(ws)
        Others = []
        for cell in ws['C'][2:]:
            Others.append(cell.value)
    except:
        print('cannot generate Other CUG Gift Subs Deductions Sheet')
        pass

    # Recharge by Denomination

    try:
        RVbD_dl = listfilter.filter_for_month(whole[16]) 
        RVBD_DF = df_converter.RVbD_list_to_df(RVbD_dl) 
        wb.create_sheet('Recharge by Denomination')
        
        ws = wb['Recharge by Denomination']

        RVBD_DF = RVBD_DF.pivot(index='Entry Date',
                            columns='Denomination',
                            values='Unique Users')

        rows = dataframe_to_rows(RVBD_DF)
        for r_idx, row in enumerate(rows,1):
            for c_idx, value in enumerate(row,1):
                ws.cell(row=r_idx,column=c_idx,value=value)

        ws = customizer.customize_rvbd(ws)
    except:
        print('cannot generate Recharge by Denomination Sheet')
        pass

    #----------

    # Summary Revenue

    wb.create_sheet('Summary Revenue',0)

    _len = len(Voice_Billable)
    table = []
    for i in range(0,_len):
        row = []
        row.append(SR_Date[i])
        row.append(Voice_Billable[i])
        row.append(SMS_Billable[i])
        row.append(GPRS_Billable[i])
        row.append(Subscriptions[i])
        row.append(Balance_Transfer[i])
        row.append(Expired_Revenue[i])
        row.append(Corporate_Plans[i])
        row.append(Others[i])
        table.append(row)

    ws = wb['Summary Revenue']

    srev_df = pd.DataFrame(table,columns=['Date','Voice Billable','SMS Billable',
                                        'GPRS Billable','Subscriptions',
                                        'Balance Transfer','Expired Revenue',
                                        'Corporate Plans','Others'])

    rows = dataframe_to_rows(srev_df)
    for r_idx, row in enumerate(rows,1):
        for c_idx, value in enumerate(row,1):
            ws.cell(row=r_idx,column=c_idx,value=value)

    _len = len(ws['B'])
    ws = customizer.customize_srev(ws)
    ws = grapher.graph_srev(ws, _len)

    #---------------------------

    del wb['Sheet']
    
    # create filename

    parent_dir = 'output\\'
    filename = generate_filename.generate_filename() + '.xlsx'
    wb.save(parent_dir + filename)
    
    return filename